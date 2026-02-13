import csv
from datetime import datetime
from io import StringIO

ALLOWED_STATUSES = {"new", "contacted", "qualified", "quoted", "accepted", "completed", "lost"}


SUPPORTED_COLUMNS = {
    "first_name",
    "last_name",
    "email",
    "phone",
    "source",
    "status",
    "language",
    "country",
    "notes",
    "next_action_date",
    "utm_source",
    "utm_medium",
    "utm_campaign",
    "utm_term",
    "utm_content",
    "gclid",
    "fbclid",
}


def _normalize_row(raw_row):
    row = {}
    for key, value in raw_row.items():
        if key is None:
            continue
        normalized_key = str(key).strip().lower()
        if normalized_key in SUPPORTED_COLUMNS:
            row[normalized_key] = str(value).strip() if value is not None else ""
    return row


def _parse_date(value):
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def map_row_to_lead_kwargs(row):
    raw_status = (row.get("status") or "new").strip().lower()
    status = raw_status if raw_status in ALLOWED_STATUSES else "new"

    return {
        "first_name": row.get("first_name") or "Unknown",
        "last_name": row.get("last_name") or "Lead",
        "email": row.get("email") or None,
        "phone": row.get("phone") or None,
        "source": row.get("source") or "import",
        "status": status,
        "stage_key": status,
        "language": row.get("language") or "it",
        "country": row.get("country") or "Italy",
        "notes": row.get("notes") or None,
        "next_action_date": _parse_date(row.get("next_action_date")),
        "utm_source": row.get("utm_source") or None,
        "utm_medium": row.get("utm_medium") or None,
        "utm_campaign": row.get("utm_campaign") or None,
        "utm_term": row.get("utm_term") or None,
        "utm_content": row.get("utm_content") or None,
        "gclid": row.get("gclid") or None,
        "fbclid": row.get("fbclid") or None,
    }


def parse_csv_leads(file_storage):
    content = file_storage.read().decode("utf-8-sig")
    reader = csv.DictReader(StringIO(content))
    return [_normalize_row(row) for row in reader]


def parse_xlsx_leads(file_storage):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX import requires openpyxl. Install it and retry.") from exc

    workbook = load_workbook(file_storage, data_only=True)
    sheet = workbook.active

    headers = []
    rows = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx == 0:
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            continue

        raw_row = {}
        for col_idx, cell in enumerate(row):
            key = headers[col_idx] if col_idx < len(headers) else ""
            if key:
                raw_row[key] = "" if cell is None else str(cell)
        rows.append(_normalize_row(raw_row))

    return rows


def parse_uploaded_leads(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".csv"):
        return parse_csv_leads(file_storage)
    if filename.endswith(".xlsx"):
        return parse_xlsx_leads(file_storage)
    raise ValueError("Unsupported file format. Use .csv or .xlsx.")
